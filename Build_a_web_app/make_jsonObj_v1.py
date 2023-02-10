from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import json

def create_student_dict(students, class_size):

    wb = load_workbook(filename=students)
    ws = wb.active

    my_list = []

    last_column = len(list(ws.columns))
    last_row = class_size

    for row in range(2, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value

        if my_dict.get("name") is None: # Skip objects with NoneType
                continue

        my_list.append(my_dict)

    data = json.dumps(my_list, sort_keys=True, indent=4, ensure_ascii=False)    
    
    return(data)